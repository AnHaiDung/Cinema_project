from openpyxl import load_workbook


path = r"C:\Users\Admin\Downloads\cgv\outputs\cinema_project_excel\Cinema-Technical_CGV_Project.xlsx"
workbook = load_workbook(path, data_only=False)

print("sheets=", workbook.sheetnames)
checks = [
    ("1. Project Charter", "B2"),
    ("1. Project Charter", "B5"),
    ("0. Roadmap", "B1"),
    ("2. Master Planxxx", "A4"),
    ("Price", "B1"),
    ("2. Price", "A1"),
    ("4. Nhân sự dự án", "A1"),
    ("7. Cập nhật tiến độ", "F2"),
]
for sheet, cell in checks:
    print(sheet, cell, workbook[sheet][cell].value)

text_parts = []
for worksheet in workbook.worksheets:
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=min(worksheet.max_row, 120),
        min_col=1,
        max_col=min(worksheet.max_column, 20),
    ):
        for cell in row:
            if cell.value is not None:
                text_parts.append(str(cell.value))

text = " ".join(text_parts)
for term in ["DỰ ÁN XXX", "CÔNG TY YYY", "SUNWORLD", "QLĐV VÀ ĐỊNH MỨC THỨC ĂN"]:
    print(term, term in text)

formula_errors = []
for worksheet in workbook.worksheets:
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and any(
                error in cell.value
                for error in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]
            ):
                formula_errors.append((worksheet.title, cell.coordinate, cell.value))
print("formula_error_like_cells", formula_errors[:20], "count=", len(formula_errors))
