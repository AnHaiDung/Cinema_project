from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill


INPUT = Path(r"C:\Users\Admin\Downloads\Cinema-Technical.xlsx")
OUTPUT_DIR = Path(r"C:\Users\Admin\Downloads\cgv\outputs\cinema_project_excel")
OUTPUT = OUTPUT_DIR / "Cinema-Technical_CGV_Project.xlsx"

PROJECT_NAME = "CGV Cinema Web Application"
PROJECT_SHORT = "CGV Cinema"
OWNER_NAME = "An Hải Dũng"
ORG = "Bộ môn CNPM"
CUSTOMER = "Dự án quản lý rạp phim CGV Cinema"
START_DATE = "08/07/2026"
END_DATE = "22/07/2026"


def norm(name):
    return "".join(name.lower().split())


def clear_area(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def set_row(ws, row_idx, values, template_row=None, start_col=1):
    template_row = template_row or row_idx
    for offset, value in enumerate(values):
        cell = ws.cell(row=row_idx, column=start_col + offset)
        if isinstance(cell, MergedCell):
            continue
        template = ws.cell(row=template_row, column=start_col + offset)
        copy_style(template, cell)
        cell.value = value


def set_title(cell, value):
    cell.value = value
    cell.font = copy(cell.font)
    cell.font = Font(name=cell.font.name or "Calibri", size=cell.font.sz or 14, bold=True, color=cell.font.color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def update_project_charter(ws):
    ws["A1"] = "PROJECT CHARTER"
    ws["A2"] = "Tên dự án:"
    ws["B2"] = PROJECT_NAME
    ws["A3"] = "Đơn vị yêu cầu:"
    ws["B3"] = CUSTOMER
    ws["A4"] = "Ban chỉ đạo dự án:"
    ws["B4"] = OWNER_NAME
    ws["A5"] = "Giám đốc dự án:"
    ws["B5"] = OWNER_NAME
    ws["A6"] = "Quản trị dự án:"
    ws["B6"] = OWNER_NAME
    ws["A7"] = "Đơn vị triển khai:"
    ws["B7"] = ORG
    ws["B8"] = START_DATE
    ws["E8"] = END_DATE
    ws["B9"] = (
        "Triển khai thành công hệ thống web đặt vé và quản lý rạp phim CGV Cinema:\n"
        "- Khách hàng xem phim, xem suất chiếu, chọn ghế và đặt vé\n"
        "- Quản trị viên quản lý phim, phòng chiếu và lịch chiếu\n"
        "- Phân quyền ROLE_ADMIN / ROLE_CUSTOMER bằng Spring Security\n"
        "- Dữ liệu vận hành qua Spring Data JPA và H2 database"
    )
    ws["B10"] = (
        "Triển khai trong phạm vi project Spring Boot hiện tại:\n"
        "- Module tài khoản: đăng ký, đăng nhập, đăng xuất\n"
        "- Module khách hàng: trang chủ, chi tiết phim, đặt vé, lịch sử vé\n"
        "- Module quản trị: phim, phòng chiếu, suất chiếu, bộ lọc lịch chiếu"
    )
    ws["E10"] = (
        "- Chưa tích hợp thanh toán online\n"
        "- Chưa có hoàn/hủy vé\n"
        "- Chưa có sơ đồ ghế động theo từng phòng\n"
        "- Chưa có báo cáo doanh thu nâng cao"
    )
    ws["E11"] = ORG
    ws["B13"] = (
        "Source code Spring Boot MVC hoàn chỉnh theo phạm vi hiện tại\n"
        "Tài liệu SRS, roadmap, ước lượng chi phí, danh sách nhân sự và nhật ký tiến độ\n"
        "Hệ thống chạy được với các chức năng đăng nhập, phân quyền, CRUD và đặt vé"
    )


def update_roadmap(ws):
    ws["B1"] = "KẾ HOẠCH CÔNG NGHỆ THÔNG TIN 2026"
    ws["B2"] = "Projects and Initiatives - CGV Cinema Web Application"
    ws["D2"] = 2026
    clear_area(ws, 4, 31, 2, 3)
    roadmap = [
        ("Khởi tạo dự án", ""),
        ("1.1", "Xác định phạm vi, actor và nghiệp vụ chính của hệ thống rạp phim"),
        ("1.2", "Lập Project Charter, roadmap, nhân sự và ước lượng chi phí"),
        ("Phân tích và thiết kế", ""),
        ("2.1", "Viết SRS cho đăng ký/đăng nhập, phân quyền, phim, phòng chiếu, lịch chiếu, đặt vé"),
        ("2.2", "Thiết kế entity: User, Movie, Auditorium, Showtime, Booking"),
        ("2.3", "Thiết kế sơ đồ Mermaid: ERD, Use Case, luồng tổng thể, trạng thái, sitemap"),
        ("Phát triển hệ thống", ""),
        ("3.1", "Cấu hình Spring Boot, Spring Security, JPA, H2 và Thymeleaf"),
        ("3.2", "Xây chức năng khách hàng: xem phim, xem suất chiếu, chọn ghế, đặt vé"),
        ("3.3", "Xây chức năng quản trị phim: thêm, sửa, xóa, upload poster"),
        ("3.4", "Xây chức năng quản trị phòng chiếu và lịch chiếu"),
        ("3.5", "Bổ sung validate custom hiển thị lỗi đỏ dưới form"),
        ("Kiểm thử và bàn giao", ""),
        ("4.1", "Test đăng ký, đăng nhập, phân quyền admin/customer"),
        ("4.2", "Test CRUD phim, phòng chiếu, lịch chiếu và lọc lịch chiếu"),
        ("4.3", "Test đặt vé, chống đặt trùng ghế và xem lịch sử vé"),
        ("4.4", "Hoàn thiện tài liệu, Excel quản lý và bàn giao source"),
    ]
    for index, (code, text) in enumerate(roadmap, start=4):
        ws.cell(index, 2).value = code
        ws.cell(index, 3).value = text
        if text == "":
            ws.cell(index, 2).font = copy(ws.cell(4, 2).font)
            ws.cell(index, 2).fill = copy(ws.cell(4, 2).fill)


def update_master_plan_xxx(ws):
    ws["A2"] = "KẾ HOẠCH TỔNG THỂ DỰ ÁN"
    ws["A4"] = "TRIỂN KHAI DỰ ÁN CGV CINEMA WEB APPLICATION"
    clear_area(ws, 8, 40, 1, 7)
    tasks = [
        ("KHỞI TẠO DỰ ÁN", "", "", "", "", "", ""),
        ("KT01", "Xác định phạm vi và actor", 1, "Done", OWNER_NAME, ORG, START_DATE),
        ("KT02", "Lập Project Charter và roadmap", 1, "Done", OWNER_NAME, ORG, START_DATE),
        ("KT03", "Chuẩn bị template SRS và Excel quản lý", 1, "Done", OWNER_NAME, ORG, "09/07/2026"),
        ("PHÂN TÍCH THIẾT KẾ", "", "", "", "", "", ""),
        ("PT01", "Phân tích nghiệp vụ đăng ký, đăng nhập và phân quyền", 1, "Done", OWNER_NAME, ORG, "09/07/2026"),
        ("PT02", "Phân tích nghiệp vụ phim, phòng chiếu, lịch chiếu", 1, "Done", OWNER_NAME, ORG, "10/07/2026"),
        ("PT03", "Phân tích nghiệp vụ đặt vé, chọn ghế, lịch sử vé", 1, "Done", OWNER_NAME, ORG, "10/07/2026"),
        ("PT04", "Thiết kế entity và quan hệ dữ liệu JPA", 1, "Done", OWNER_NAME, ORG, "11/07/2026"),
        ("PHÁT TRIỂN", "", "", "", "", "", ""),
        ("DEV01", "Cấu hình Spring Boot, Security, JPA, H2", 1, "Done", OWNER_NAME, ORG, "11/07/2026"),
        ("DEV02", "Xây chức năng đăng ký, đăng nhập, đăng xuất", 1, "Done", OWNER_NAME, ORG, "12/07/2026"),
        ("DEV03", "Xây trang chủ, chi tiết phim và danh sách suất chiếu", 1, "Done", OWNER_NAME, ORG, "13/07/2026"),
        ("DEV04", "Xây chọn ghế, đặt vé và xem vé đã đặt", 2, "Done", OWNER_NAME, ORG, "14/07/2026"),
        ("DEV05", "Xây CRUD phim, upload poster", 1, "Done", OWNER_NAME, ORG, "15/07/2026"),
        ("DEV06", "Xây CRUD phòng chiếu", 1, "Done", OWNER_NAME, ORG, "15/07/2026"),
        ("DEV07", "Xây CRUD và lọc lịch chiếu", 1, "Done", OWNER_NAME, ORG, "16/07/2026"),
        ("DEV08", "Sửa validate custom thay cho HTML5 validation", 1, "Done", OWNER_NAME, ORG, "21/07/2026"),
        ("KIỂM THỬ - BÀN GIAO", "", "", "", "", "", ""),
        ("QA01", "Kiểm thử phân quyền admin/customer", 1, "In progress", OWNER_NAME, ORG, "21/07/2026"),
        ("QA02", "Kiểm thử CRUD và đặt vé", 1, "In progress", OWNER_NAME, ORG, "21/07/2026"),
        ("DOC01", "Hoàn thiện SRS, Excel kỹ thuật và tài liệu bàn giao", 1, "In progress", OWNER_NAME, ORG, "22/07/2026"),
    ]
    for idx, row in enumerate(tasks, start=8):
        set_row(ws, idx, row, template_row=9)


def update_price(ws):
    ws["B1"] = "PRICE / COST ESTIMATE - CGV CINEMA PROJECT"
    rows = [
        (1, "Viết SRS cho CGV Cinema", "PM/BA", 1, 800000, "=D5*E5", "Đặc tả yêu cầu, actor, use case, sơ đồ Mermaid"),
        (2, "Lập roadmap và Project Charter", "PM", 1, 800000, "=D6*E6", "Kế hoạch phạm vi, thời gian và bàn giao"),
        (3, "Đăng nhập / đăng ký / phân quyền", "Developer", 1, 900000, "=D7*E7", "Spring Security, ROLE_ADMIN, ROLE_CUSTOMER"),
        (4, "Khách hàng xem phim, suất chiếu, đặt vé", "Developer", 2, 900000, "=D8*E8", "Trang chủ, chi tiết phim, chọn ghế, booking"),
        (5, "Quản lý phim và poster", "Developer", 1, 900000, "=D9*E9", "CRUD Movie, upload poster"),
        (6, "Quản lý phòng chiếu", "Developer", 1, 900000, "=D10*E10", "CRUD Auditorium"),
        (7, "Quản lý lịch chiếu", "Developer", 1, 900000, "=D11*E11", "CRUD Showtime, trạng thái suất chiếu"),
        (8, "Lọc lịch chiếu", "Developer", 1, 900000, "=D12*E12", "Lọc theo ngày, phim, phòng"),
        (9, "Giao diện Thymeleaf + CSS", "Frontend Dev", 1, 800000, "=D13*E13", "Layout, form, table, seat map"),
        (10, "Validate custom", "Developer", 1, 700000, "=D14*E14", "Thông báo lỗi đỏ dưới field, không dùng HTML5 popup"),
        (11, "Kiểm thử tổng hợp", "QA", 1, 700000, "=D15*E15", "Test login, CRUD, booking, validate"),
        (12, "Tài liệu + bàn giao", "PM/Dev", 1, 800000, "=D16*E16", "SRS, Excel kỹ thuật, hướng dẫn chạy"),
    ]
    for idx, row in enumerate(rows, start=5):
        set_row(ws, idx, row, template_row=5)
    ws["B18"] = "Tổng chi phí dự kiến"
    ws["F18"] = "=SUM(F5:F16)"
    ws["G18"] = "VNĐ"


def update_price_2(ws):
    ws["A1"] = PROJECT_NAME
    ws["B4"] = f"Công ty: {ORG}"
    ws["B5"] = f"Dự án: {PROJECT_NAME}"
    ws["F5"] = f"Từ ngày : {START_DATE}"
    ws["B6"] = f"Khách hàng: {CUSTOMER}"
    ws["F6"] = f"Đến ngày: {END_DATE}"
    clear_area(ws, 11, 40, 1, 6)
    rows = [
        ("I", "PHÂN TÍCH THIẾT KẾ", None, None, None, None),
        (1, "Lập kế hoạch dự án", "Project Charter, roadmap, nhân sự, chi phí", 1, 1, "Hoàn thành"),
        (2, "Phân tích yêu cầu nghiệp vụ", "Đăng ký, đăng nhập, phim, phòng chiếu, lịch chiếu, đặt vé", 2, 2, "Hoàn thành"),
        (3, "Thiết kế kiến trúc hệ thống", "Spring Boot MVC, Thymeleaf, Spring Security, JPA, H2", 1, 1, "Hoàn thành"),
        (4, "Thiết kế CSDL", "User, Movie, Auditorium, Showtime, Booking", 1, 1, "Hoàn thành"),
        ("II", "PHÁT TRIỂN VÀ XÂY DỰNG", None, None, None, None),
        (1, "Đăng nhập/đăng ký/phân quyền", "ROLE_ADMIN, ROLE_CUSTOMER", 1, 1, "Hoàn thành"),
        (2, "Khách hàng xem phim và đặt vé", "Trang chủ, chi tiết phim, chọn ghế, booking confirmation", 2, 2, "Hoàn thành"),
        (3, "Quản lý phim", "CRUD phim, upload poster", 1, 1, "Hoàn thành"),
        (4, "Quản lý phòng chiếu", "CRUD phòng chiếu, kiểm tra trùng tên", 1, 1, "Hoàn thành"),
        (5, "Quản lý lịch chiếu", "CRUD, trạng thái, lọc ngày/phim/phòng", 1, 1, "Hoàn thành"),
        (6, "Validate custom", "Không dùng HTML5 validation, hiển thị lỗi đỏ dưới field", 1, 1, "Hoàn thành"),
        ("III", "KIỂM THỬ VÀ BÀN GIAO", None, None, None, None),
        (1, "Kiểm thử chức năng", "Login, register, CRUD, booking, validate", 1, 1, "Đang thực hiện"),
        (2, "Hoàn thiện tài liệu", "SRS, Excel kỹ thuật, hướng dẫn bàn giao", 1, 1, "Đang thực hiện"),
    ]
    for idx, row in enumerate(rows, start=11):
        set_row(ws, idx, row, template_row=12)


def update_staff(ws):
    ws["A1"] = "DANH SÁCH NHÂN SỰ THAM GIA DỰ ÁN CGV CINEMA"
    clear_area(ws, 3, 15, 1, 8)
    rows = [
        ("Bộ môn CNPM", None, None, None, None, None, None, None),
        (1, OWNER_NAME, ORG, "Giám đốc dự án", "Quản lý phạm vi, tiến độ, chất lượng", None, "", PROJECT_NAME),
        (2, OWNER_NAME, ORG, "BA", "Phân tích nghiệp vụ và viết SRS", None, "", PROJECT_NAME),
        (3, OWNER_NAME, ORG, "Backend Developer", "Spring Boot, Security, JPA, service/repository", None, "", PROJECT_NAME),
        (4, OWNER_NAME, ORG, "Frontend Developer", "Thymeleaf template, CSS, form, validate", None, "", PROJECT_NAME),
        (5, OWNER_NAME, ORG, "Tester", "Kiểm thử chức năng, validate và phân quyền", None, "", PROJECT_NAME),
        ("Khách hàng / Người dùng", None, None, None, None, None, None, None),
        (1, "Quản trị viên rạp", CUSTOMER, "Người dùng nghiệp vụ", "Quản lý phim, phòng chiếu, lịch chiếu", None, "", None),
        (2, "Khách hàng đặt vé", CUSTOMER, "Người dùng cuối", "Xem phim, chọn ghế, đặt vé, xem vé đã đặt", None, "", None),
    ]
    for idx, row in enumerate(rows, start=3):
        set_row(ws, idx, row, template_row=4)


def update_progress(ws):
    ws["F2"] = PROJECT_NAME
    ws["F3"] = "Action Tracking Sheet\nNhật ký hoạt động dự án CGV Cinema"
    ws["F5"] = f"Project Manager: {OWNER_NAME}\nQuản trị dự án"
    ws["F6"] = f"Workstream Lead: {OWNER_NAME}\nTrưởng nhóm công việc"
    ws["F7"] = f"Updated by: {OWNER_NAME}\nCập nhật bởi"
    ws["F8"] = "Updated on: 21/07/2026\nNgày cập nhật"
    clear_area(ws, 13, 24, 1, 11)
    rows = [
        (1, "08/07/2026", "Project", "Khởi tạo", "Project Charter", "Xác định phạm vi CGV Cinema và kế hoạch triển khai", "Completed", "08/07/2026", None, OWNER_NAME, None),
        (2, "09/07/2026", "Project", "SRS", "Đặc tả nghiệp vụ", "Viết SRS cho tài khoản, phim, phòng chiếu, lịch chiếu, đặt vé", "Completed", "10/07/2026", None, OWNER_NAME, None),
        (3, "11/07/2026", "Source code", "Backend", "Entity và repository", "Hoàn thiện User, Movie, Auditorium, Showtime, Booking", "Completed", "12/07/2026", None, OWNER_NAME, None),
        (4, "13/07/2026", "Source code", "Frontend", "Trang khách hàng", "Trang chủ, chi tiết phim, booking, lịch sử vé", "Completed", "14/07/2026", None, OWNER_NAME, None),
        (5, "15/07/2026", "Source code", "Admin", "CRUD quản trị", "Quản lý phim, phòng chiếu, lịch chiếu", "Completed", "16/07/2026", None, OWNER_NAME, None),
        (6, "21/07/2026", "Yêu cầu chỉnh sửa", "Frontend", "Validate form", "Bỏ HTML5 validation và hiện một dòng lỗi đỏ bên dưới field", "Completed", "21/07/2026", None, OWNER_NAME, None),
        (7, "21/07/2026", "Yêu cầu chỉnh sửa", "Tài liệu", "Excel kỹ thuật", "Chuyển workbook template sang project CGV Cinema, loại sheet 3. Master Plan", "In Progress", "22/07/2026", None, OWNER_NAME, None),
    ]
    for idx, row in enumerate(rows, start=13):
        set_row(ws, idx, row, template_row=13)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(INPUT)

    for sheet_name in list(workbook.sheetnames):
        if norm(sheet_name) == norm("3. Master Plan"):
            workbook.remove(workbook[sheet_name])

    update_roadmap(workbook["0. Roadmap"])
    update_project_charter(workbook["1. Project Charter"])
    update_master_plan_xxx(workbook["2. Master Planxxx"])
    update_price(workbook["Price"])
    update_price_2(workbook["2. Price"])
    update_staff(workbook["4. Nhân sự dự án"])
    update_progress(workbook["7. Cập nhật tiến độ"])

    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "\n" in cell.value:
                    cell.alignment = copy(cell.alignment)
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=True,
                    )

    workbook.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False, read_only=True)
    print(check.sheetnames)
    if any(norm(name) == norm("3. Master Plan") for name in check.sheetnames):
        raise SystemExit("Sheet 3. Master Plan vẫn còn trong workbook.")
    print(f"SAVED {OUTPUT}")


if __name__ == "__main__":
    main()
