from openpyxl import load_workbook


workbook = load_workbook(r"C:\Users\Admin\Downloads\Cinema-Technical.xlsx", data_only=False)
print(workbook.sheetnames)
for worksheet in workbook.worksheets:
    print("SHEET", worksheet.title, worksheet.max_row, worksheet.max_column)
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=min(worksheet.max_row, 20),
        max_col=min(worksheet.max_column, 10),
        values_only=True,
    ):
        print(row)
    print("---")
