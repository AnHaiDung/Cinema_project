from pathlib import Path

from openpyxl import load_workbook


INPUT = Path(r"C:\Users\Admin\Downloads\Cinema-Technical.xlsx")
OUTPUT_DIR = Path(r"C:\Users\Admin\Downloads\cgv\outputs\excel_without_masterplan")
OUTPUT = OUTPUT_DIR / "Cinema-Technical_without_masterplan.xlsx"
TARGET_SHEET = "3. masterplan"


def normalize_sheet_name(name):
    return "".join(name.lower().split())


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(INPUT)
    print("BEFORE")
    for sheet_name in workbook.sheetnames:
        print(sheet_name)

    matches = [
        name
        for name in workbook.sheetnames
        if normalize_sheet_name(name) == normalize_sheet_name(TARGET_SHEET)
    ]
    if not matches:
        raise SystemExit(f'Không tìm thấy sheet "{TARGET_SHEET}".')

    workbook.remove(workbook[matches[0]])
    workbook.save(OUTPUT)

    check = load_workbook(OUTPUT, read_only=True, data_only=False)
    print("AFTER")
    for sheet_name in check.sheetnames:
        print(sheet_name)
    if matches[0] in check.sheetnames:
        raise SystemExit(f'Xoá sheet "{matches[0]}" chưa thành công.')
    print(f"SAVED {OUTPUT}")


if __name__ == "__main__":
    main()
